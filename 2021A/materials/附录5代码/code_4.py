import openpyxl
import code_1
import numpy
import sympy

workbook = openpyxl.load_workbook(filename="附件4.xlsx")
sheet_1 = workbook['理想抛物面顶点坐标']
sheet_2 = workbook['调整后主索节点编号及坐标']
sheet_3 = workbook['促动器顶端伸缩量']

variation = -0.60

alpha0 = 36.795
beta0 = 78.169
alpha = alpha0 / 180 * numpy.pi
beta = beta0 / 180 * numpy.pi
cp_length = (1 - 0.466) * 300
A = -numpy.cos(beta) * numpy.cos(alpha)
B = -numpy.cos(beta) * numpy.sin(alpha)
C = -numpy.sin(beta)
px = cp_length * A
py = cp_length * B
pz = cp_length * C
Dp = -(A * px + B * py + C * pz)
ox = (300 - variation) * A
oy = (300 - variation) * B
oz = (300 - variation) * C
Do = -(A * ox + B * oy + C * oz)
D = 2 * Do - Dp

# 填入理想抛物面的顶点坐标
sheet_1.cell(row=2, column=1).value = ox
sheet_1.cell(row=2, column=2).value = oy
sheet_1.cell(row=2, column=3).value = oz

move_node_name_list = []
offset_list = {}

for j in range(2226):
    name = code_1.node_namelist[j]
    node = code_1.main_cable_node[name]
    x = node.x
    y = node.y
    z = node.z
    direction = code_1.node_information[name].direction_1
    delta_x = px - x
    delta_y = py - y
    delta_z = pz - z
    distance = numpy.linalg.norm(numpy.cross(numpy.array([delta_x, delta_y, delta_z]), numpy.array([A, B, C])))
    if distance < 150:
        t = sympy.Symbol('t')
        X = x + t * direction[0]
        Y = y + t * direction[1]
        Z = z + t * direction[2]
        eq = [(X - px) ** 2 + (Y - py) ** 2 + (Z - pz) ** 2 - (A * X + B * Y + C * Z + D) ** 2]
        ans = sympy.solve(eq, t)
        if len(ans) == 2:
            if abs(ans[0][0]) < abs(ans[1][0]):
                offset = ans[0][0]
            else:
                offset = ans[1][0]
        else:
            offset = ans[t]
        move_node_name_list.append(name)
        offset_list[name] = offset

line_num = 2
for k in move_node_name_list:
    node = code_1.main_cable_node[k]
    information = code_1.node_information[k]
    direction = information.direction_1
    x = node.x + direction[0] * offset_list[k]
    y = node.y + direction[1] * offset_list[k]
    z = node.z + direction[2] * offset_list[k]
    sheet_2['A%d' % line_num].value = k
    sheet_2['B%d' % line_num].value = float(x)
    sheet_2['C%d' % line_num].value = float(y)
    sheet_2['D%d' % line_num].value = float(z)
    sheet_3['A%d' % line_num].value = k
    sheet_3['B%d' % line_num].value = float(offset_list[k])
    line_num = line_num + 1

workbook.save(filename="附件4.xlsx")
workbook.close()
